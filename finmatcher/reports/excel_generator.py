"""
Excel report generator for FinMatcher v2.0 Enterprise Upgrade.

This module generates three types of Excel reports:
1. Matched records report ({Statement_File_Name}_record.xlsx)
2. Unmatched financial emails report (Other_receipts_email.xlsx)
3. Unmatched non-financial emails report (unmatch_email_records.xlsx)

All reports include proper formatting (currency with 2 decimals, ISO 8601 dates)
and support append-or-create logic for idempotent operations.

Validates Requirements: 9.1, 9.2, 9.3, 9.4, 9.5, 9.6, 9.7, 9.8
"""

import os
import logging
import re
from pathlib import Path
from typing import List
from datetime import date
from decimal import Decimal
import openpyxl
from openpyxl.styles import numbers

from finmatcher.storage.models import MatchResult, Receipt, MatchConfidence

logger = logging.getLogger(__name__)

# Regex pattern for illegal characters in Excel (control characters except tab, newline, carriage return)
ILLEGAL_CHARACTERS_RE = re.compile(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]')


class ExcelReportGenerator:
    """
    Generates Excel reports with matched and unmatched records.
    
    Validates Requirements: 9.1, 9.2, 9.3, 9.4, 9.5, 9.6, 9.7, 9.8
    """
    
    def __init__(self, output_dir: str):
        """
        Initialize report generator.
        
        Args:
            output_dir: Directory for temporary report files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"Excel report generator initialized with output dir: {output_dir}")
    
    def generate_matched_report(self, matches: List[MatchResult], 
                               statement_name: str) -> str:
        """
        Generate report for matched transactions (Exact or High confidence).
        
        Args:
            matches: List of MatchResult objects with Exact or High confidence
            statement_name: Name of statement file
            
        Returns:
            Path to generated Excel file
            
        Validates Requirements:
        - 9.1: Write Exact/High confidence matches to {Statement_File_Name}_record.xlsx
        - 9.4: Include all required columns
        - 9.5: Append rows without overwriting existing data
        - 9.6: Include header row with column names
        - 9.7: Format currency with two decimal places
        - 9.8: Format dates using ISO 8601 (YYYY-MM-DD)
        """
        filename = f"{statement_name}_record.xlsx"
        file_path = self.output_dir / filename
        
        # Check if file exists
        if file_path.exists():
            # Load existing workbook and append
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            logger.info(f"Appending to existing matched report: {filename}")
        else:
            # Create new workbook with headers
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Matched Records"
            
            # Add header row (Requirement 9.6)
            headers = [
                "Transaction Date",
                "Transaction Amount",
                "Transaction Description",
                "Receipt Date",
                "Receipt Amount",
                "Receipt Description",
                "Amount Score",
                "Date Score",
                "Semantic Score",
                "Composite Score",
                "Confidence"
            ]
            ws.append(headers)
            logger.info(f"Created new matched report: {filename}")
        
        # Add matched records (Requirement 9.4)
        for match in matches:
            row = [
                self._format_date(match.transaction.transaction_date),  # ISO 8601
                self._format_currency(match.transaction.amount),  # 2 decimals
                self._sanitize_text(match.transaction.description),
                self._format_date(match.receipt.receipt_date),  # ISO 8601
                self._format_currency(match.receipt.amount) if match.receipt.amount else "",
                self._sanitize_text(match.receipt.description),
                round(match.amount_score, 2),
                round(match.date_score, 2),
                round(match.semantic_score, 2),
                round(match.composite_score, 2),
                match.confidence.value
            ]
            ws.append(row)
        
        # Apply formatting to currency and date columns
        self._apply_column_formatting(ws)
        
        # Save workbook
        wb.save(file_path)
        logger.info(f"Saved matched report with {len(matches)} records: {file_path}")
        
        return str(file_path)
    
    def generate_unmatched_financial_report(self, receipts: List[Receipt]) -> str:
        """
        Generate report for unmatched financial emails (Low confidence, is_financial=True).
        
        Args:
            receipts: List of Receipt objects with Low confidence and is_financial=True
            
        Returns:
            Path to generated Excel file
            
        Validates Requirements:
        - 9.2: Write Low confidence financial receipts to Other_receipts_email.xlsx
        - 9.4: Include required columns
        - 9.5: Append rows without overwriting existing data
        - 9.6: Include header row with column names
        - 9.7: Format currency with two decimal places
        - 9.8: Format dates using ISO 8601 (YYYY-MM-DD)
        """
        filename = "Other_receipts_email.xlsx"
        file_path = self.output_dir / filename
        
        # Check if file exists
        if file_path.exists():
            # Load existing workbook and append
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            logger.info(f"Appending to existing unmatched financial report: {filename}")
        else:
            # Create new workbook with headers
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Unmatched Financial"
            
            # Add header row
            headers = [
                "Receipt Date",
                "Receipt Amount",
                "Receipt Description",
                "Source",
                "Filter Method"
            ]
            ws.append(headers)
            logger.info(f"Created new unmatched financial report: {filename}")
        
        # Add unmatched financial records
        for receipt in receipts:
            row = [
                self._format_date(receipt.receipt_date),  # ISO 8601
                self._format_currency(receipt.amount) if receipt.amount else "",  # 2 decimals
                self._sanitize_text(receipt.description),
                receipt.source,
                receipt.filter_method.value if receipt.filter_method else ""
            ]
            ws.append(row)
        
        # Apply formatting
        self._apply_column_formatting(ws)
        
        # Save workbook
        wb.save(file_path)
        logger.info(f"Saved unmatched financial report with {len(receipts)} records: {file_path}")
        
        return str(file_path)
    
    def generate_unmatched_nonfinancial_report(self, receipts: List[Receipt]) -> str:
        """
        Generate report for non-financial emails (Low confidence, is_financial=False).
        
        Args:
            receipts: List of Receipt objects with Low confidence and is_financial=False
            
        Returns:
            Path to generated Excel file
            
        Validates Requirements:
        - 9.3: Write Low confidence non-financial receipts to unmatch_email_records.xlsx
        - 9.4: Include required columns
        - 9.5: Append rows without overwriting existing data
        - 9.6: Include header row with column names
        - 9.8: Format dates using ISO 8601 (YYYY-MM-DD)
        """
        filename = "unmatch_email_records.xlsx"
        file_path = self.output_dir / filename
        
        # Check if file exists
        if file_path.exists():
            # Load existing workbook and append
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            logger.info(f"Appending to existing non-financial report: {filename}")
        else:
            # Create new workbook with headers
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Non-Financial Emails"
            
            # Add header row
            headers = [
                "Receipt Date",
                "Receipt Description",
                "Source",
                "Filter Method"
            ]
            ws.append(headers)
            logger.info(f"Created new non-financial report: {filename}")
        
        # Add non-financial records
        for receipt in receipts:
            row = [
                self._format_date(receipt.receipt_date),  # ISO 8601
                self._sanitize_text(receipt.description),
                receipt.source,
                receipt.filter_method.value if receipt.filter_method else ""
            ]
            ws.append(row)
        
        # Apply formatting
        self._apply_column_formatting(ws)
        
        # Save workbook
        wb.save(file_path)
        logger.info(f"Saved non-financial report with {len(receipts)} records: {file_path}")
        
        return str(file_path)
    
    def _format_currency(self, amount: Decimal) -> float:
        """
        Format currency with 2 decimal places.
        
        Args:
            amount: Decimal amount
            
        Returns:
            Float rounded to 2 decimal places
            
        Validates Requirement: 9.7 - Format currency columns with two decimal places
        """
        if amount is None:
            return 0.0
        return round(float(amount), 2)
    
    def _format_date(self, date_obj: date) -> str:
        """
        Format date as ISO 8601 (YYYY-MM-DD).
        
        Args:
            date_obj: Date object
            
        Returns:
            ISO 8601 formatted date string
            
        Validates Requirement: 9.8 - Format date columns using ISO 8601 format
        """
        if date_obj is None:
            return ""
        return date_obj.strftime("%Y-%m-%d")
    
    def _sanitize_text(self, text: str) -> str:
        """
        Sanitize text by removing illegal characters that Excel doesn't support.
        
        Excel/openpyxl doesn't allow control characters (0x00-0x1F except tab, newline, CR)
        and some other control characters (0x7F-0x9F).
        
        Args:
            text: Input text string
            
        Returns:
            Sanitized text with illegal characters removed, or "[empty]" if result is empty
        """
        if text is None:
            return "[empty]"
        if not isinstance(text, str):
            text = str(text)
        # Remove illegal characters
        sanitized = ILLEGAL_CHARACTERS_RE.sub('', text)
        # If sanitization results in empty string, return placeholder
        if not sanitized or sanitized.isspace():
            return "[empty]"
        return sanitized
    
    def _apply_column_formatting(self, ws):
        """
        Apply number formatting to currency and date columns.
        
        Args:
            ws: Worksheet object
        """
        # Apply currency format to amount columns (2 decimal places)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                # Check if cell contains a number (amount columns)
                if isinstance(cell.value, (int, float)):
                    # Apply currency format with 2 decimals
                    cell.number_format = '0.00'
