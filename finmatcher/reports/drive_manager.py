"""
Google Drive manager for FinMatcher v2.0 Enterprise Upgrade.

This module handles Google Drive folder hierarchy management and file uploads
with idempotent operations, caching, and retry logic.

8-Folder Hierarchy:
- FinMatcher_Excel_Reports/ (root)
  ├── {Statement_File_Name}_record.xlsx/
  ├── Other_receipts_email.xlsx/
  ├── unmatch_email_records.xlsx/
  ├── Attach_files/
  ├── Attach_Image/
  ├── Unmatch_Email_Attach_files/
  └── unmatch_attach_image/

Validates Requirements: 3.1, 3.2, 3.3, 3.4, 3.5, 3.6, 3.7, 3.8, 3.9, 3.10, 4.1, 4.2, 4.3, 4.4, 4.5, 4.6, 10.3
"""

import os
import time
import logging
from pathlib import Path
from typing import Optional, Dict
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from googleapiclient.errors import HttpError
import openpyxl
from io import BytesIO

from finmatcher.storage.models import FolderStructure, AttachmentType, MatchConfidence

logger = logging.getLogger(__name__)


class DriveManager:
    """
    Manages Google Drive folder hierarchy and file uploads with idempotent operations.
    
    Validates Requirements: 3.1-3.10, 4.1-4.6, 10.3
    """
    
    SCOPES = ['https://www.googleapis.com/auth/drive.file']
    
    def __init__(self, credentials_path: str, root_folder_name: str = "FinMatcher_Excel_Reports"):
        """
        Initialize Google Drive API client.
        
        Args:
            credentials_path: Path to credentials.json file
            root_folder_name: Name of root folder
            
        Validates Requirement: 3.1 - Create or retrieve root folder
        """
        self.credentials_path = credentials_path
        self.root_folder_name = root_folder_name
        self.service = None
        self.folder_cache: Dict[str, str] = {}
        self.folder_structure: Optional[FolderStructure] = None
        
        # Authenticate and build service
        self._authenticate()
        
        logger.info(f"Drive manager initialized for root folder: {root_folder_name}")
    
    def _authenticate(self):
        """
        Authenticate with Google Drive API using OAuth 2.0.
        """
        creds = None
        token_path = self.credentials_path.replace('credentials.json', 'token.json')
        
        # Load existing token
        if os.path.exists(token_path):
            creds = Credentials.from_authorized_user_file(token_path, self.SCOPES)
        
        # Refresh or obtain new credentials
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    self.credentials_path, self.SCOPES
                )
                creds = flow.run_local_server(port=0)
            
            # Save credentials
            with open(token_path, 'w') as token:
                token.write(creds.to_json())
        
        self.service = build('drive', 'v3', credentials=creds)
        logger.info("Google Drive authentication successful")
    
    def initialize_folder_structure(self) -> FolderStructure:
        """
        Create or retrieve complete 8-folder hierarchy.
        
        Returns:
            FolderStructure object containing all folder IDs
            
        Validates Requirements: 3.1, 3.2, 3.3, 3.4, 3.5, 3.6, 3.7, 3.8
        """
        # Get or create root folder
        root_id = self.get_or_create_folder(self.root_folder_name)
        
        # Create standard subfolders
        other_receipts_id = self.get_or_create_folder("Other_receipts_email.xlsx", root_id)
        unmatch_email_id = self.get_or_create_folder("unmatch_email_records.xlsx", root_id)
        attach_files_id = self.get_or_create_folder("Attach_files", root_id)
        attach_image_id = self.get_or_create_folder("Attach_Image", root_id)
        unmatch_attach_files_id = self.get_or_create_folder("Unmatch_Email_Attach_files", root_id)
        unmatch_attach_image_id = self.get_or_create_folder("unmatch_attach_image", root_id)
        
        self.folder_structure = FolderStructure(
            root_id=root_id,
            statement_folders={},
            other_receipts_id=other_receipts_id,
            unmatch_email_id=unmatch_email_id,
            attach_files_id=attach_files_id,
            attach_image_id=attach_image_id,
            unmatch_attach_files_id=unmatch_attach_files_id,
            unmatch_attach_image_id=unmatch_attach_image_id
        )
        
        logger.info("Folder structure initialized successfully")
        return self.folder_structure
    
    def get_or_create_statement_folder(self, statement_name: str) -> str:
        """
        Get or create folder for statement-specific records.
        
        Args:
            statement_name: Name of statement file
            
        Returns:
            Folder ID
            
        Validates Requirement: 3.2 - Create subfolder for each statement file
        """
        if not self.folder_structure:
            raise RuntimeError("Folder structure not initialized")
        
        # Generate folder name: {Statement_File_Name}_record.xlsx
        folder_name = f"{statement_name}_record.xlsx"
        
        # Check cache
        if folder_name in self.folder_structure.statement_folders:
            return self.folder_structure.statement_folders[folder_name]
        
        # Create folder
        folder_id = self.get_or_create_folder(folder_name, self.folder_structure.root_id)
        self.folder_structure.statement_folders[folder_name] = folder_id
        
        return folder_id
    
    def get_or_create_folder(self, name: str, parent_id: Optional[str] = None) -> str:
        """
        Get existing folder or create if not exists (idempotent).
        
        Args:
            name: Folder name
            parent_id: Parent folder ID (None for root)
            
        Returns:
            Folder ID
            
        Validates Requirements:
        - 3.9: Query Google Drive metadata before attempting creation
        - 3.10: Return existing folder identifier if folder already exists
        """
        # Check cache
        cache_key = f"{parent_id}:{name}"
        if cache_key in self.folder_cache:
            return self.folder_cache[cache_key]
        
        # Query existing folders
        query = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
        if parent_id:
            query += f" and '{parent_id}' in parents"
        
        try:
            results = self.service.files().list(
                q=query,
                spaces='drive',
                fields='files(id, name)'
            ).execute()
            
            files = results.get('files', [])
            
            if files:
                # Folder exists, return ID
                folder_id = files[0]['id']
                logger.debug(f"Found existing folder: {name} (ID: {folder_id})")
                self.folder_cache[cache_key] = folder_id
                return folder_id
            else:
                # Create new folder
                file_metadata = {
                    'name': name,
                    'mimeType': 'application/vnd.google-apps.folder'
                }
                if parent_id:
                    file_metadata['parents'] = [parent_id]
                
                folder = self.service.files().create(
                    body=file_metadata,
                    fields='id'
                ).execute()
                
                folder_id = folder['id']
                logger.info(f"Created new folder: {name} (ID: {folder_id})")
                self.folder_cache[cache_key] = folder_id
                return folder_id
                
        except HttpError as e:
            logger.error(f"Error creating/retrieving folder {name}: {e}")
            raise
    
    def upload_excel_report(self, file_path: str, folder_id: str, append: bool = True) -> str:
        """
        Upload Excel report with append-or-create logic.
        
        Args:
            file_path: Local file path
            folder_id: Target Drive folder ID
            append: If True and file exists, append rows; else create new
            
        Returns:
            File ID
            
        Validates Requirements:
        - 4.1: Check if file exists in target folder
        - 4.2: Append new rows to existing file if exists
        - 4.3: Create new file if doesn't exist
        """
        filename = os.path.basename(file_path)
        
        # Check if file exists
        query = f"name='{filename}' and '{folder_id}' in parents and trashed=false"
        
        try:
            results = self.service.files().list(
                q=query,
                spaces='drive',
                fields='files(id, name)'
            ).execute()
            
            files = results.get('files', [])
            
            if files and append:
                # File exists, append rows
                existing_id = files[0]['id']
                logger.info(f"Appending to existing file: {filename}")
                
                # Download existing file
                request = self.service.files().get_media(fileId=existing_id)
                existing_data = BytesIO(request.execute())
                
                # Load both workbooks
                existing_wb = openpyxl.load_workbook(existing_data)
                new_wb = openpyxl.load_workbook(file_path)
                
                # Append rows from new to existing
                existing_ws = existing_wb.active
                new_ws = new_wb.active
                
                # Skip header row in new file
                for row in new_ws.iter_rows(min_row=2, values_only=True):
                    existing_ws.append(row)
                
                # Save combined workbook
                combined_data = BytesIO()
                existing_wb.save(combined_data)
                combined_data.seek(0)
                
                # Update file in Drive
                media = MediaFileUpload(
                    file_path,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    resumable=True
                )
                
                updated_file = self.service.files().update(
                    fileId=existing_id,
                    media_body=media
                ).execute()
                
                logger.info(f"File updated successfully: {filename}")
                return existing_id
            else:
                # Create new file
                file_metadata = {
                    'name': filename,
                    'parents': [folder_id]
                }
                
                media = MediaFileUpload(
                    file_path,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    resumable=True
                )
                
                file = self.service.files().create(
                    body=file_metadata,
                    media_body=media,
                    fields='id'
                ).execute()
                
                file_id = file['id']
                logger.info(f"File created successfully: {filename} (ID: {file_id})")
                
                # Verify upload (Requirement 4.5)
                if self._verify_upload(file_id):
                    return file_id
                else:
                    raise RuntimeError(f"Upload verification failed for {filename}")
                    
        except HttpError as e:
            logger.error(f"Error uploading Excel report {filename}: {e}")
            raise
    
    def upload_attachment(self, file_path: str, folder_id: str, source_prefix: str) -> str:
        """
        Upload attachment with unique filename.
        
        Args:
            file_path: Local file path
            folder_id: Target Drive folder ID
            source_prefix: Prefix for filename (e.g., "email", "statement")
            
        Returns:
            File ID
            
        Validates Requirements:
        - 4.4: Generate unique filename using pattern {Source}_{Counter}.{Extension}
        - 4.5: Verify successful upload
        - 4.6: Retry up to 3 times with exponential backoff
        - 10.3: Log file name, target folder, and upload timestamp
        """
        original_name = os.path.basename(file_path)
        unique_filename = self._generate_unique_filename(original_name, source_prefix, folder_id)
        
        # Retry logic with exponential backoff
        max_retries = 3
        base_delay = 1.0
        
        for attempt in range(max_retries):
            try:
                file_metadata = {
                    'name': unique_filename,
                    'parents': [folder_id]
                }
                
                # Determine MIME type
                mime_type = self._get_mime_type(file_path)
                
                media = MediaFileUpload(
                    file_path,
                    mimetype=mime_type,
                    resumable=True
                )
                
                file = self.service.files().create(
                    body=file_metadata,
                    media_body=media,
                    fields='id'
                ).execute()
                
                file_id = file['id']
                
                # Verify upload
                if self._verify_upload(file_id):
                    logger.info(
                        f"Attachment uploaded: {unique_filename} "
                        f"(folder: {folder_id}, file_id: {file_id})"
                    )
                    return file_id
                else:
                    if attempt < max_retries - 1:
                        logger.warning(f"Upload verification failed, retry {attempt + 1}")
                        continue
                    else:
                        raise RuntimeError(f"Upload verification failed after {max_retries} retries")
                        
            except HttpError as e:
                if attempt < max_retries - 1:
                    delay = base_delay * (2 ** attempt)
                    logger.warning(f"Upload failed: {e}, retrying in {delay}s")
                    time.sleep(delay)
                else:
                    logger.error(f"Upload failed after {max_retries} retries: {e}")
                    raise
        
        raise RuntimeError(f"Failed to upload attachment after {max_retries} retries")
    
    def _verify_upload(self, file_id: str) -> bool:
        """
        Verify file exists in Drive by querying metadata.
        
        Args:
            file_id: File ID to verify
            
        Returns:
            True if file exists, False otherwise
            
        Validates Requirement: 4.5 - Verify successful upload by querying file metadata
        """
        try:
            file = self.service.files().get(
                fileId=file_id,
                fields='id, name'
            ).execute()
            
            return file is not None
            
        except HttpError as e:
            logger.error(f"Verification failed for file {file_id}: {e}")
            return False
    
    def _generate_unique_filename(self, original_name: str, prefix: str, folder_id: str) -> str:
        """
        Generate unique filename: {prefix}_{counter}.{extension}
        
        Args:
            original_name: Original filename
            prefix: Prefix for filename
            folder_id: Target folder ID
            
        Returns:
            Unique filename
            
        Validates Requirement: 4.4 - Generate unique filename pattern
        """
        # Extract extension
        name_parts = original_name.rsplit('.', 1)
        extension = name_parts[1] if len(name_parts) > 1 else ''
        
        # Find next available counter
        counter = 1
        while True:
            if extension:
                filename = f"{prefix}_{counter}.{extension}"
            else:
                filename = f"{prefix}_{counter}"
            
            # Check if filename exists in folder
            query = f"name='{filename}' and '{folder_id}' in parents and trashed=false"
            results = self.service.files().list(
                q=query,
                spaces='drive',
                fields='files(id)'
            ).execute()
            
            if not results.get('files', []):
                return filename
            
            counter += 1
    
    def _get_mime_type(self, file_path: str) -> str:
        """Get MIME type based on file extension."""
        extension = os.path.splitext(file_path)[1].lower()
        
        mime_types = {
            '.pdf': 'application/pdf',
            '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.gif': 'image/gif'
        }
        
        return mime_types.get(extension, 'application/octet-stream')
    
    def classify_attachment(self, file_path: str) -> AttachmentType:
        """
        Classify attachment as DOCUMENT or IMAGE based on file extension.
        
        Args:
            file_path: Path to attachment file
            
        Returns:
            AttachmentType.DOCUMENT or AttachmentType.IMAGE
            
        Validates Requirements:
        - 5.1: Process receipts with attachments
        - 5.2: Classify PDF, DOC, DOCX as Attachment_File
        - 5.3: Classify JPG, JPEG, PNG, GIF as Attachment_Image
        """
        extension = os.path.splitext(file_path)[1].lower()
        
        # Document extensions
        document_extensions = {'.pdf', '.doc', '.docx'}
        # Image extensions
        image_extensions = {'.jpg', '.jpeg', '.png', '.gif'}
        
        if extension in document_extensions:
            return AttachmentType.DOCUMENT
        elif extension in image_extensions:
            return AttachmentType.IMAGE
        else:
            # Default to document for unknown types
            logger.warning(f"Unknown file extension {extension}, defaulting to DOCUMENT")
            return AttachmentType.DOCUMENT
    
    def route_attachment(self, file_path: str, confidence: MatchConfidence, 
                        source_prefix: str = "email") -> str:
        """
        Route attachment to appropriate Drive folder based on confidence and file type.
        
        Args:
            file_path: Path to attachment file
            confidence: Match confidence level
            source_prefix: Prefix for filename (default: "email")
            
        Returns:
            File ID of uploaded attachment
            
        Validates Requirements:
        - 5.4: Route matched documents to "Attach_files"
        - 5.5: Route matched images to "Attach_Image"
        - 5.6: Route unmatched documents to "Unmatch_Email_Attach_files"
        - 5.7: Route unmatched images to "unmatch_attach_image"
        """
        if not self.folder_structure:
            raise RuntimeError("Folder structure not initialized")
        
        # Classify attachment type
        attachment_type = self.classify_attachment(file_path)
        
        # Determine target folder based on confidence and type
        if confidence in (MatchConfidence.EXACT, MatchConfidence.HIGH):
            # Matched attachments
            if attachment_type == AttachmentType.DOCUMENT:
                folder_id = self.folder_structure.attach_files_id
                logger.debug(f"Routing matched document to Attach_files")
            else:  # IMAGE
                folder_id = self.folder_structure.attach_image_id
                logger.debug(f"Routing matched image to Attach_Image")
        else:  # LOW confidence
            # Unmatched attachments
            if attachment_type == AttachmentType.DOCUMENT:
                folder_id = self.folder_structure.unmatch_attach_files_id
                logger.debug(f"Routing unmatched document to Unmatch_Email_Attach_files")
            else:  # IMAGE
                folder_id = self.folder_structure.unmatch_attach_image_id
                logger.debug(f"Routing unmatched image to unmatch_attach_image")
        
        # Upload attachment to determined folder
        return self.upload_attachment(file_path, folder_id, source_prefix)
