"""
Property-based tests for Excel Report Schema.

This module contains property-based tests using Hypothesis to verify
that all generated Excel reports contain the complete set of required columns.

Testing Framework: pytest + Hypothesis
Feature: finmatcher-v2-upgrade
Task: 11.3 - Write property test for report schema
Property 16: Report Schema Contains All Required Columns
"""

import pytest
from hypothesis import given, strategies as st, settings
from decimal import Decimal
from datetime import date
from typing import List
import tempfile
import os
import openpyxl

from finmatcher.reports.excel_generator import ExcelReportGenerator
from finmatcher.storage.models import (
    MatchResult, Transaction, Receipt, MatchConfidence, FilterMethod
)


# Configure Hypothesis settings for FinMatcher
# Using max_examples=20 for faster test execution as specified in task 11.3
settings.register_profile("finmatcher", max_examples=20)
settings.load_profile("finmatcher")


# Required columns as specified in Requirement 9.4
REQUIRED_MATCHED_COLUMNS = [
    "Transaction Date",
    "Transaction Amount",
    "Transaction Description",
    "Receipt Date",
    "Receipt Amount",
    "Receipt Description",
    "Amount Score",
    "Date Score",
    "Semantic Score",
    "Composite Score"
]


# Custom strategies for generating test data
@st.composite
def transaction_strategy(draw):
    """Generate a valid Transaction object."""
    return Transaction(
        id=draw(st.integers(min_value=1, max_value=10000)),
        statement_file=draw(st.text(min_size=1, max_size=50, alphabet=st.characters(whitelist_categories=('Lu', 'Ll', 'Nd', 'P')))),
        transaction_date=draw(st.dates(min_value=date(2020, 1, 1), max_value=date(2025, 12, 31))),
        amount=draw(st.decimals(min_value=Decimal("0.01"), max_value=Decimal("10000.00"), places=2)),
        description=draw(st.text(min_size=1, max_size=100))
    )


@st.composite
def receipt_strategy(draw):
    """Generate a valid Receipt object."""
    return Receipt(
        id=draw(st.integers(min_value=1, max_value=10000)),
        source=draw(st.sampled_from(["email", "manual"])),
        receipt_date=draw(st.dates(min_value=date(2020, 1, 1), max_value=date(2025, 12, 31))),
        amount=draw(st.one_of(
            st.none(),
            st.decimals(min_value=Decimal("0.01"), max_value=Decimal("10000.00"), places=2)
        )),
        description=draw(st.text(min_size=1, max_size=100)),
        is_financial=draw(st.booleans()),
        filter_method=draw(st.one_of(
            st.none(),
            st.sampled_from([FilterMethod.AUTO_ACCEPT, FilterMethod.AUTO_REJECT, FilterMethod.AI_VERIFIED])
        )),
        attachments=[]
    )


@st.composite
def match_result_strategy(draw):
    """Generate a valid MatchResult object."""
    return MatchResult(
        transaction=draw(transaction_strategy()),
        receipt=draw(receipt_strategy()),
        amount_score=draw(st.floats(min_value=0.0, max_value=1.0)),
        date_score=draw(st.floats(min_value=0.0, max_value=1.0)),
        semantic_score=draw(st.floats(min_value=0.0, max_value=1.0)),
        composite_score=draw(st.floats(min_value=0.0, max_value=1.0)),
        confidence=draw(st.sampled_from([MatchConfidence.EXACT, MatchConfidence.HIGH, MatchConfidence.LOW]))
    )


class TestReportSchemaProperties:
    """
    Property-based tests for Excel Report Schema.
    
    Tests verify that all generated reports contain the required columns
    as specified in Requirement 9.4.
    """
    
    # Feature: finmatcher-v2-upgrade, Property 16: Report Schema Contains All Required Columns
    @given(
        matches=st.lists(match_result_strategy(), min_size=1, max_size=20),
        statement_name=st.text(min_size=1, max_size=50, alphabet=st.characters(whitelist_categories=('Lu', 'Ll', 'Nd')))
    )
    @settings(max_examples=10)
    def test_matched_report_contains_all_required_columns(
        self,
        matches: List[MatchResult],
        statement_name: str
    ):
        """
        **Validates: Requirements 9.4**
        
        Property 16: Report Schema Contains All Required Columns
        
        For any generated Excel report for matched transactions, it must contain
        all required columns: transaction date, amount, description, receipt date,
        receipt amount, receipt description, Amount_Score, Date_Score, Semantic_Score,
        and Composite_Score.
        
        This property verifies that:
        1. All matched reports include the complete set of required columns
        2. Column headers are present in the first row
        3. The column order and naming match the specification
        4. The schema is consistent across all generated reports
        """
        # Create temporary directory for test reports
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create Excel report generator
            generator = ExcelReportGenerator(output_dir=temp_dir)
            
            # Generate matched report
            report_path = generator.generate_matched_report(matches, statement_name)
            
            # Verify file was created
            assert os.path.exists(report_path), (
                f"Report file was not created at {report_path}"
            )
            
            # Load the workbook and get the active sheet
            wb = openpyxl.load_workbook(report_path)
            ws = wb.active
            
            # Get header row (first row)
            header_row = ws[1]
            actual_headers = [cell.value for cell in header_row if cell.value is not None]
            
            # Assert: All required columns must be present
            for required_column in REQUIRED_MATCHED_COLUMNS:
                assert required_column in actual_headers, (
                    f"Required column '{required_column}' is missing from report. "
                    f"Found headers: {actual_headers}"
                )
            
            # Assert: Headers must be in the first row
            assert ws.max_row >= 1, "Report must have at least a header row"
            
            # Assert: Number of data rows matches number of matches
            expected_rows = len(matches) + 1  # +1 for header
            assert ws.max_row == expected_rows, (
                f"Report should have {expected_rows} rows (1 header + {len(matches)} data), "
                f"but has {ws.max_row} rows"
            )
            
            # Close workbook
            wb.close()
    
    @given(
        matches=st.lists(match_result_strategy(), min_size=1, max_size=20),
        statement_name=st.text(min_size=1, max_size=50, alphabet=st.characters(whitelist_categories=('Lu', 'Ll', 'Nd')))
    )
    @settings(max_examples=10)
    def test_matched_report_column_order_is_correct(
        self,
        matches: List[MatchResult],
        statement_name: str
    ):
        """
        **Validates: Requirements 9.4**
        
        Property 16: Report Schema Contains All Required Columns (Column Order)
        
        For any generated Excel report, the columns must appear in the correct order
        as specified in the requirements.
        
        This property verifies that:
        1. Transaction columns appear before receipt columns
        2. Score columns appear after transaction and receipt columns
        3. The column order is consistent with the specification
        """
        # Create temporary directory for test reports
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create Excel report generator
            generator = ExcelReportGenerator(output_dir=temp_dir)
            
            # Generate matched report
            report_path = generator.generate_matched_report(matches, statement_name)
            
            # Load the workbook and get the active sheet
            wb = openpyxl.load_workbook(report_path)
            ws = wb.active
            
            # Get header row
            header_row = ws[1]
            actual_headers = [cell.value for cell in header_row if cell.value is not None]
            
            # Define expected column order
            expected_order = [
                "Transaction Date",
                "Transaction Amount",
                "Transaction Description",
                "Receipt Date",
                "Receipt Amount",
                "Receipt Description",
                "Amount Score",
                "Date Score",
                "Semantic Score",
                "Composite Score"
            ]
            
            # Assert: All expected columns are present in the correct order
            for i, expected_col in enumerate(expected_order):
                assert i < len(actual_headers), (
                    f"Expected column '{expected_col}' at position {i}, "
                    f"but only {len(actual_headers)} columns found"
                )
                assert actual_headers[i] == expected_col, (
                    f"Column at position {i} should be '{expected_col}', "
                    f"but found '{actual_headers[i]}'"
                )
            
            # Close workbook
            wb.close()
    
    @given(
        matches=st.lists(match_result_strategy(), min_size=1, max_size=20),
        statement_name=st.text(min_size=1, max_size=50, alphabet=st.characters(whitelist_categories=('Lu', 'Ll', 'Nd')))
    )
    @settings(max_examples=10)
    def test_matched_report_data_rows_match_schema(
        self,
        matches: List[MatchResult],
        statement_name: str
    ):
        """
        **Validates: Requirements 9.4**
        
        Property 16: Report Schema Contains All Required Columns (Data Integrity)
        
        For any generated Excel report, all data rows must contain values
        in the correct columns matching the schema.
        
        This property verifies that:
        1. Each data row has values in all required columns
        2. Transaction data appears in transaction columns
        3. Receipt data appears in receipt columns
        4. Score data appears in score columns
        """
        # Create temporary directory for test reports
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create Excel report generator
            generator = ExcelReportGenerator(output_dir=temp_dir)
            
            # Generate matched report
            report_path = generator.generate_matched_report(matches, statement_name)
            
            # Load the workbook and get the active sheet
            wb = openpyxl.load_workbook(report_path)
            ws = wb.active
            
            # Get header row to map column names to indices
            header_row = ws[1]
            headers = [cell.value for cell in header_row if cell.value is not None]
            
            # Create column index mapping
            col_indices = {header: idx for idx, header in enumerate(headers)}
            
            # Verify each data row (skip header row)
            for row_idx, match in enumerate(matches, start=2):
                row = ws[row_idx]
                
                # Assert: Transaction Date column has a value
                txn_date_col = col_indices.get("Transaction Date")
                assert txn_date_col is not None, "Transaction Date column not found"
                assert row[txn_date_col].value is not None, (
                    f"Row {row_idx}: Transaction Date should not be empty"
                )
                
                # Assert: Transaction Amount column has a value
                txn_amount_col = col_indices.get("Transaction Amount")
                assert txn_amount_col is not None, "Transaction Amount column not found"
                assert row[txn_amount_col].value is not None, (
                    f"Row {row_idx}: Transaction Amount should not be empty"
                )
                
                # Assert: Transaction Description column has a value
                txn_desc_col = col_indices.get("Transaction Description")
                assert txn_desc_col is not None, "Transaction Description column not found"
                assert row[txn_desc_col].value is not None, (
                    f"Row {row_idx}: Transaction Description should not be empty"
                )
                
                # Assert: Receipt Date column has a value
                rec_date_col = col_indices.get("Receipt Date")
                assert rec_date_col is not None, "Receipt Date column not found"
                assert row[rec_date_col].value is not None, (
                    f"Row {row_idx}: Receipt Date should not be empty"
                )
                
                # Assert: Score columns have numeric values
                for score_col in ["Amount Score", "Date Score", "Semantic Score", "Composite Score"]:
                    score_col_idx = col_indices.get(score_col)
                    assert score_col_idx is not None, f"{score_col} column not found"
                    score_value = row[score_col_idx].value
                    assert score_value is not None, (
                        f"Row {row_idx}: {score_col} should not be empty"
                    )
                    assert isinstance(score_value, (int, float)), (
                        f"Row {row_idx}: {score_col} should be numeric, got {type(score_value)}"
                    )
            
            # Close workbook
            wb.close()
    
    @given(
        receipts=st.lists(receipt_strategy(), min_size=1, max_size=20)
    )
    @settings(max_examples=10)
    def test_unmatched_financial_report_contains_required_columns(
        self,
        receipts: List[Receipt]
    ):
        """
        **Validates: Requirements 9.4**
        
        Property 16: Report Schema Contains All Required Columns (Unmatched Financial)
        
        For any generated Excel report for unmatched financial emails,
        it must contain the required columns for receipt data.
        
        This property verifies that:
        1. Unmatched financial reports include receipt-specific columns
        2. Column headers are present in the first row
        3. The schema is appropriate for unmatched receipts
        """
        # Create temporary directory for test reports
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create Excel report generator
            generator = ExcelReportGenerator(output_dir=temp_dir)
            
            # Generate unmatched financial report
            report_path = generator.generate_unmatched_financial_report(receipts)
            
            # Verify file was created
            assert os.path.exists(report_path), (
                f"Report file was not created at {report_path}"
            )
            
            # Load the workbook and get the active sheet
            wb = openpyxl.load_workbook(report_path)
            ws = wb.active
            
            # Get header row
            header_row = ws[1]
            actual_headers = [cell.value for cell in header_row if cell.value is not None]
            
            # Required columns for unmatched financial report
            required_columns = [
                "Receipt Date",
                "Receipt Amount",
                "Receipt Description",
                "Source",
                "Filter Method"
            ]
            
            # Assert: All required columns must be present
            for required_column in required_columns:
                assert required_column in actual_headers, (
                    f"Required column '{required_column}' is missing from unmatched financial report. "
                    f"Found headers: {actual_headers}"
                )
            
            # Assert: Number of data rows matches number of receipts
            expected_rows = len(receipts) + 1  # +1 for header
            assert ws.max_row == expected_rows, (
                f"Report should have {expected_rows} rows (1 header + {len(receipts)} data), "
                f"but has {ws.max_row} rows"
            )
            
            # Close workbook
            wb.close()
    
    @given(
        receipts=st.lists(receipt_strategy(), min_size=1, max_size=20)
    )
    @settings(max_examples=10)
    def test_unmatched_nonfinancial_report_contains_required_columns(
        self,
        receipts: List[Receipt]
    ):
        """
        **Validates: Requirements 9.4**
        
        Property 16: Report Schema Contains All Required Columns (Unmatched Non-Financial)
        
        For any generated Excel report for non-financial emails,
        it must contain the required columns for non-financial data.
        
        This property verifies that:
        1. Non-financial reports include appropriate columns
        2. Column headers are present in the first row
        3. The schema is appropriate for non-financial emails
        """
        # Create temporary directory for test reports
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create Excel report generator
            generator = ExcelReportGenerator(output_dir=temp_dir)
            
            # Generate unmatched non-financial report
            report_path = generator.generate_unmatched_nonfinancial_report(receipts)
            
            # Verify file was created
            assert os.path.exists(report_path), (
                f"Report file was not created at {report_path}"
            )
            
            # Load the workbook and get the active sheet
            wb = openpyxl.load_workbook(report_path)
            ws = wb.active
            
            # Get header row
            header_row = ws[1]
            actual_headers = [cell.value for cell in header_row if cell.value is not None]
            
            # Required columns for non-financial report
            required_columns = [
                "Receipt Date",
                "Receipt Description",
                "Source",
                "Filter Method"
            ]
            
            # Assert: All required columns must be present
            for required_column in required_columns:
                assert required_column in actual_headers, (
                    f"Required column '{required_column}' is missing from non-financial report. "
                    f"Found headers: {actual_headers}"
                )
            
            # Assert: Number of data rows matches number of receipts
            expected_rows = len(receipts) + 1  # +1 for header
            assert ws.max_row == expected_rows, (
                f"Report should have {expected_rows} rows (1 header + {len(receipts)} data), "
                f"but has {ws.max_row} rows"
            )
            
            # Close workbook
            wb.close()
    
    @given(
        matches_batch1=st.lists(match_result_strategy(), min_size=1, max_size=10),
        matches_batch2=st.lists(match_result_strategy(), min_size=1, max_size=10),
        statement_name=st.text(min_size=1, max_size=50, alphabet=st.characters(whitelist_categories=('Lu', 'Ll', 'Nd')))
    )
    @settings(max_examples=10)
    def test_appended_report_maintains_schema(
        self,
        matches_batch1: List[MatchResult],
        matches_batch2: List[MatchResult],
        statement_name: str
    ):
        """
        **Validates: Requirements 9.4, 9.5**
        
        Property 16: Report Schema Contains All Required Columns (Append Operations)
        
        For any Excel report that is appended to (idempotent operation),
        the schema must remain consistent and all required columns must be present.
        
        This property verifies that:
        1. Appending data preserves the original schema
        2. All required columns remain present after append
        3. The header row is not duplicated
        4. Data integrity is maintained across append operations
        """
        # Create temporary directory for test reports
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create Excel report generator
            generator = ExcelReportGenerator(output_dir=temp_dir)
            
            # Generate initial report with first batch
            report_path = generator.generate_matched_report(matches_batch1, statement_name)
            
            # Append second batch to the same report
            report_path = generator.generate_matched_report(matches_batch2, statement_name)
            
            # Load the workbook and get the active sheet
            wb = openpyxl.load_workbook(report_path)
            ws = wb.active
            
            # Get header row
            header_row = ws[1]
            actual_headers = [cell.value for cell in header_row if cell.value is not None]
            
            # Assert: All required columns must still be present
            for required_column in REQUIRED_MATCHED_COLUMNS:
                assert required_column in actual_headers, (
                    f"Required column '{required_column}' is missing after append. "
                    f"Found headers: {actual_headers}"
                )
            
            # Assert: Total rows should be header + both batches
            expected_rows = 1 + len(matches_batch1) + len(matches_batch2)
            assert ws.max_row == expected_rows, (
                f"Report should have {expected_rows} rows after append "
                f"(1 header + {len(matches_batch1)} + {len(matches_batch2)} data), "
                f"but has {ws.max_row} rows"
            )
            
            # Assert: Only one header row exists (no duplication)
            # Check that row 2 contains data, not headers
            if ws.max_row > 1:
                row2 = ws[2]
                row2_values = [cell.value for cell in row2]
                # First cell should be a date string, not "Transaction Date"
                assert row2_values[0] != "Transaction Date", (
                    "Header row appears to be duplicated at row 2"
                )
            
            # Close workbook
            wb.close()

