"""
Simplified End-to-End Integration Test for FinMatcher v2.0.

This test validates the core workflow components work together correctly
without complex dependencies that might cause timeouts.

Testing Framework: pytest
Feature: finmatcher-v2-upgrade
Task: 19.1 - Write end-to-end integration test
"""

import pytest
import tempfile
import os
import shutil
from datetime import date
from decimal import Decimal
from unittest.mock import Mock

from finmatcher.core.financial_filter import FinancialFilter, FilterMethod
from finmatcher.storage.database_manager import DatabaseManager
from finmatcher.storage.checkpoint_manager import CheckpointManager
from finmatcher.storage.models import Transaction, Receipt
from finmatcher.reports.excel_generator import ExcelReportGenerator


@pytest.fixture
def temp_dir():
    """Create a temporary directory for test files."""
    temp_path = tempfile.mkdtemp()
    yield temp_path
    shutil.rmtree(temp_path, ignore_errors=True)


@pytest.fixture
def temp_db(temp_dir):
    """Create a temporary database for testing."""
    db_path = os.path.join(temp_dir, 'test_finmatcher.db')
    yield db_path


@pytest.fixture
def db_manager(temp_db):
    """Create a DatabaseManager instance."""
    manager = DatabaseManager(temp_db, pool_size=5)
    manager.initialize_schema()
    yield manager
    manager.close()


class TestSimplifiedE2EWorkflow:
    """
    Simplified end-to-end integration tests.
    
    Tests core workflow components:
    1. Email filtering (financial vs non-financial)
    2. Database persistence (transactions and receipts)
    3. Checkpoint and resume capability
    4. Report generation
    """
    
    def test_email_filtering_workflow(self):
        """
        Test email filtering accepts financial and rejects non-financial emails.
        
        Validates:
        - Financial keywords trigger auto-accept
        - Marketing keywords trigger auto-reject
        - Filter methods are tracked correctly
        """
        # Create mock DeepSeek client
        mock_client = Mock()
        mock_client.verify_financial_email = Mock(return_value=True)
        
        financial_filter = FinancialFilter(deepseek_client=mock_client)
        
        # Test financial email (should be accepted)
        financial_email = {
            'subject': 'Your invoice #12345',
            'sender': 'billing@company.com',
            'body': 'Invoice for services'
        }
        
        result = financial_filter.filter_email(financial_email)
        assert result is not None, "Financial email should be accepted"
        assert result['is_financial'] == True
        assert result['filter_method'] == FilterMethod.AUTO_ACCEPT.value
        
        # Test marketing email (should be rejected)
        marketing_email = {
            'subject': 'Unsubscribe from our newsletter',
            'sender': 'marketing@spam.com',
            'body': 'Click here to unsubscribe'
        }
        
        result = financial_filter.filter_email(marketing_email)
        assert result is None, "Marketing email should be rejected"
        
        # Verify statistics
        stats = financial_filter.get_statistics()
        assert stats['total_processed'] == 2
        assert stats['auto_accepted'] == 1
        assert stats['auto_rejected'] == 1
    
    def test_database_persistence_workflow(self, db_manager):
        """
        Test database persistence for transactions and receipts.
        
        Validates:
        - Transactions can be saved and retrieved
        - Receipts can be saved and retrieved
        - Data integrity is maintained
        """
        # Create and save transactions
        transactions = [
            Transaction(
                id=None,
                statement_file="test_statement.csv",
                transaction_date=date(2024, 1, 15),
                amount=Decimal("100.00"),
                description="Test transaction 1"
            ),
            Transaction(
                id=None,
                statement_file="test_statement.csv",
                transaction_date=date(2024, 1, 16),
                amount=Decimal("50.00"),
                description="Test transaction 2"
            )
        ]
        
        saved_ids = []
        for txn in transactions:
            txn_id = db_manager.save_transaction(txn)
            saved_ids.append(txn_id)
            assert txn_id > 0, "Transaction should be saved with valid ID"
        
        # Retrieve transactions
        retrieved_txns = db_manager.get_transactions(limit=10)
        assert len(retrieved_txns) == 2, "Should retrieve 2 transactions"
        assert retrieved_txns[0].amount == Decimal("100.00")
        assert retrieved_txns[1].amount == Decimal("50.00")
        
        # Create and save receipts
        receipts = [
            Receipt(
                id=None,
                source='email',
                receipt_date=date(2024, 1, 15),
                amount=Decimal("100.00"),
                description="Test receipt 1",
                is_financial=True,
                filter_method=FilterMethod.AUTO_ACCEPT,
                attachments=[]
            ),
            Receipt(
                id=None,
                source='email',
                receipt_date=date(2024, 1, 16),
                amount=Decimal("50.00"),
                description="Test receipt 2",
                is_financial=True,
                filter_method=FilterMethod.AUTO_ACCEPT,
                attachments=[]
            )
        ]
        
        for receipt in receipts:
            receipt_id = db_manager.save_receipt(receipt)
            assert receipt_id > 0, "Receipt should be saved with valid ID"
        
        # Retrieve receipts
        retrieved_receipts = db_manager.get_receipts(limit=10)
        assert len(retrieved_receipts) == 2, "Should retrieve 2 receipts"
        assert retrieved_receipts[0].amount == Decimal("100.00")
        assert retrieved_receipts[1].amount == Decimal("50.00")
    
    def test_checkpoint_resume_workflow(self, db_manager):
        """
        Test checkpoint and resume capability.
        
        Validates:
        - Checkpoints are saved correctly
        - Checkpoints can be loaded after restart
        - Checkpoints can be deleted on completion
        """
        # Create checkpoint manager
        checkpoint_manager = CheckpointManager(db_manager, checkpoint_interval=10)
        
        # Save checkpoint
        checkpoint_manager.save_checkpoint(
            last_transaction_id=100,
            last_receipt_id=50
        )
        
        # Load checkpoint
        checkpoint = checkpoint_manager.load_checkpoint()
        assert checkpoint is not None, "Checkpoint should exist"
        assert checkpoint.last_transaction_id == 100
        assert checkpoint.last_receipt_id == 50
        
        # Simulate restart - create new checkpoint manager
        new_checkpoint_manager = CheckpointManager(db_manager, checkpoint_interval=10)
        resume_checkpoint = new_checkpoint_manager.load_checkpoint()
        
        assert resume_checkpoint is not None, "Should load checkpoint after restart"
        assert resume_checkpoint.last_transaction_id == 100
        assert resume_checkpoint.last_receipt_id == 50
        
        # Delete checkpoint on completion
        new_checkpoint_manager.delete_checkpoint()
        
        # Verify checkpoint is deleted
        final_checkpoint = new_checkpoint_manager.load_checkpoint()
        assert final_checkpoint is None, "Checkpoint should be deleted"
    
    def test_report_generation_workflow(self, temp_dir, db_manager):
        """
        Test Excel report generation.
        
        Validates:
        - Reports can be generated from match data
        - Report files are created correctly
        - Report contains expected data
        """
        # Create test data
        transaction = Transaction(
            id=1,
            statement_file="test_statement",
            transaction_date=date(2024, 1, 15),
            amount=Decimal("100.00"),
            description="Test transaction"
        )
        
        receipt = Receipt(
            id=1,
            source='email',
            receipt_date=date(2024, 1, 15),
            amount=Decimal("100.00"),
            description="Test receipt",
            is_financial=True,
            filter_method=FilterMethod.AUTO_ACCEPT,
            attachments=[]
        )
        
        # Create match result (manually construct to avoid matching engine)
        from finmatcher.storage.models import MatchResult, MatchConfidence
        
        match = MatchResult(
            transaction=transaction,
            receipt=receipt,
            amount_score=1.0,
            date_score=1.0,
            semantic_score=0.9,
            composite_score=0.97,
            confidence=MatchConfidence.HIGH
        )
        
        # Generate report
        output_dir = os.path.join(temp_dir, 'reports')
        os.makedirs(output_dir, exist_ok=True)
        
        excel_generator = ExcelReportGenerator(output_dir=output_dir)
        report_path = excel_generator.generate_matched_report(
            [match],
            "test_statement"
        )
        
        # Verify report was created
        assert os.path.exists(report_path), "Report should be created"
        assert report_path.endswith('.xlsx'), "Report should be Excel file"
        
        # Verify report contains data
        import openpyxl
        wb = openpyxl.load_workbook(report_path)
        ws = wb.active
        
        # Should have header row + 1 data row
        assert ws.max_row >= 2, "Report should have header + data row"
        
        # Verify header exists
        header_row = [cell.value for cell in ws[1]]
        assert 'Transaction Date' in header_row or any('Date' in str(h) for h in header_row if h)
    
    def test_complete_workflow_integration(self, temp_dir, db_manager):
        """
        Test complete workflow: filter → persist → checkpoint → report.
        
        Validates all components work together:
        1. Email filtering
        2. Database persistence
        3. Checkpoint management
        4. Report generation
        """
        # Step 1: Filter emails
        mock_client = Mock()
        mock_client.verify_financial_email = Mock(return_value=True)
        financial_filter = FinancialFilter(deepseek_client=mock_client)
        
        test_emails = [
            {'subject': 'Invoice #123', 'sender': 'billing@co.com', 'body': 'Invoice'},
            {'subject': 'Unsubscribe', 'sender': 'spam@co.com', 'body': 'Click here'},
            {'subject': 'Receipt #456', 'sender': 'store@co.com', 'body': 'Receipt'}
        ]
        
        filtered_emails = []
        for email in test_emails:
            result = financial_filter.filter_email(email)
            if result:
                filtered_emails.append(result)
        
        assert len(filtered_emails) == 2, "Should filter 2 financial emails"
        
        # Step 2: Persist to database
        transactions = []
        for i, email in enumerate(filtered_emails):
            txn = Transaction(
                id=None,
                statement_file="test_statement",
                transaction_date=date(2024, 1, 15 + i),
                amount=Decimal(f"{100 + i * 10}.00"),
                description=email['subject']
            )
            txn.id = db_manager.save_transaction(txn)
            transactions.append(txn)
        
        assert len(transactions) == 2, "Should save 2 transactions"
        
        # Step 3: Save checkpoint
        checkpoint_manager = CheckpointManager(db_manager, checkpoint_interval=1)
        checkpoint_manager.save_checkpoint(
            last_transaction_id=transactions[-1].id,
            last_receipt_id=0
        )
        
        checkpoint = checkpoint_manager.load_checkpoint()
        assert checkpoint is not None, "Checkpoint should be saved"
        
        # Step 4: Generate report (simplified - just verify it can be created)
        output_dir = os.path.join(temp_dir, 'reports')
        os.makedirs(output_dir, exist_ok=True)
        
        excel_generator = ExcelReportGenerator(output_dir=output_dir)
        
        # Create a simple match for reporting
        from finmatcher.storage.models import MatchResult, MatchConfidence
        
        receipt = Receipt(
            id=1,
            source='email',
            receipt_date=date(2024, 1, 15),
            amount=Decimal("100.00"),
            description="Test receipt",
            is_financial=True,
            filter_method=FilterMethod.AUTO_ACCEPT,
            attachments=[]
        )
        
        match = MatchResult(
            transaction=transactions[0],
            receipt=receipt,
            amount_score=1.0,
            date_score=1.0,
            semantic_score=0.9,
            composite_score=0.97,
            confidence=MatchConfidence.HIGH
        )
        
        report_path = excel_generator.generate_matched_report([match], "test_statement")
        assert os.path.exists(report_path), "Report should be generated"
        
        # Step 5: Clean up checkpoint
        checkpoint_manager.delete_checkpoint()
        assert checkpoint_manager.load_checkpoint() is None, "Checkpoint should be deleted"
        
        # Verify complete workflow succeeded
        assert len(filtered_emails) == 2, "Filtering succeeded"
        assert len(transactions) == 2, "Persistence succeeded"
        assert os.path.exists(report_path), "Report generation succeeded"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
